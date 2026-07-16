"""Excel 导出测试。"""

from __future__ import annotations

import openpyxl
import pytest

from mee.modules.medical_extractor.engine import export_rows_to_excel


def test_export_maps_columns_and_meta(fixtures_dir, tmp_path):
    template = fixtures_dir / "template.xlsx"
    out = tmp_path / "结果.xlsx"
    rows = [
        {"来源文件": "a.md", "姓名": "张三", "住院号": "T001", "主诉": "头痛", "年龄": 30,
         "_source": "张三", "_status": "success"},
        {"来源文件": "b.md", "姓名": "李四", "住院号": "T002", "主诉": "-1", "年龄": -1,
         "_source": "李四", "_status": "failed", "_error_message": "解析失败"},
    ]
    path = export_rows_to_excel(rows, str(template), str(out))
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    header = [c.value for c in ws[1]]
    assert "姓名" in header
    # 追加了元信息列
    assert "提取状态" in header and "数据来源" in header and "错误信息" in header

    name_col = header.index("姓名") + 1
    assert ws.cell(2, name_col).value == "张三"
    assert ws.cell(3, name_col).value == "李四"

    status_col = header.index("提取状态") + 1
    assert ws.cell(2, status_col).value == "success"
    assert ws.cell(3, status_col).value == "failed"


def test_export_empty_raises(fixtures_dir, tmp_path):
    with pytest.raises(ValueError):
        export_rows_to_excel([], str(fixtures_dir / "template.xlsx"), str(tmp_path / "x.xlsx"))
