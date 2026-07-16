"""全链路端到端（mock OCR + LLM）：预处理→OCR→合并→抽取→导出。

不发真实网络请求：AsyncOCRClient.process_file 与 MedicalExtractionEngine 均被替换。
验证选完文件夹后一路跑到 结果.xlsx，且各病人一行、列值正确。
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl

from mee.controllers.pipeline_controller import PipelineConfig, PipelineWorker


def _fake_ocr_results(text: str):
    """构造 AsyncOCRClient.process_file 的返回结构（含 markdown 文本）。"""
    return [{"layoutParsingResults": [{"markdown": {"text": text, "images": {}}, "outputImages": {}}]}]


def test_full_pipeline_mocked(fixtures_dir, tmp_path):
    raw = fixtures_dir / "raw"
    pre = tmp_path / "pre"
    ocr_out = tmp_path / "ocr"
    template = fixtures_dir / "template_config.json"
    output = tmp_path / "结果.xlsx"

    cfg = PipelineConfig(
        scenario="image",
        raw_input=str(raw),
        preprocess_output=str(pre),
        ocr_output=str(ocr_out),
        api_url="http://ocr",
        api_token="tok",
        ocr_model="m",
        ocr_preset="original",
        file_extensions=[".jpg"],
        enable_payment_ocr=False,
        payment_pattern="-缴费情况.jpg",
        cleanup_target="",
        cleanup_pattern="*右表格_0.md",
        selected_steps=["preprocess", "ocr_batch", "merge", "extract", "export"],
        extraction_template=str(template),
        output_excel=str(output),
        extract_llm={"provider": "DeepSeek", "api_key": "k", "model": "m"},
        make_docx=False,
    )

    worker = PipelineWorker(cfg)
    events = {}
    worker.step_completed.connect(lambda k, s, m: events.__setitem__(k, s))
    finished = {}
    worker.finished.connect(lambda ok, msg: finished.update(success=ok, message=msg))

    # mock OCR：每个文件返回一段合成文本
    fake_client = MagicMock()
    fake_client.process_file.side_effect = lambda p: _fake_ocr_results(f"OCR文本 {Path(p).stem}")

    # mock 抽取引擎：从 source（病人目录名）返回一行
    fake_engine = MagicMock()
    fake_engine.extract.side_effect = lambda content, source="": {
        "姓名": source, "住院号": f"ID-{source}", "_source": source, "_status": "success",
    }

    with patch("mee.controllers.pipeline_controller.AsyncOCRClient", return_value=fake_client), \
         patch("mee.controllers.pipeline_controller.MedicalExtractionEngine", return_value=fake_engine):
        worker.run()

    # 所有步骤成功
    for step in ("preprocess", "ocr_batch", "merge", "extract", "export"):
        assert events[step] == "success", f"{step} -> {events.get(step)}"
    assert finished["success"] is True

    # 预处理产物存在
    assert list(pre.rglob("*.jpg"))
    # 每个病人一份 merged.md
    assert (ocr_out / "张三" / "张三_merged.md").exists()
    assert (ocr_out / "李四" / "李四_merged.md").exists()

    # 结果 Excel：两行病人数据
    assert output.exists()
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    header = [c.value for c in ws[1]]
    name_col = header.index("姓名") + 1
    names = {ws.cell(r, name_col).value for r in (2, 3)}
    assert names == {"张三", "李四"}
