"""医疗病历结构化抽取引擎（纯逻辑，无 UI 依赖）。

从原 ``Medical_Excel_Agent_Pro.py`` 剥离出来，供主流水线与独立 GUI 共用：

- 多供应商 API 客户端（DeepSeek / OpenAI 兼容 / Claude），OpenAI 兼容分支
  可覆盖智谱、通义、Azure、以及任意自定义 base_url。
- ``MedicalExtractionEngine``：把一段病历文本抽取为一行结构化数据。
- ``export_rows_to_excel``：把多行结果按模板表头写入 Excel（含状态着色）。
- ``load_template_config``：从预设 JSON 或 Excel 模板表头构造字段配置。
"""

from __future__ import annotations

import json
import logging
import time
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import requests

logger = logging.getLogger(__name__)


# ==================== API 客户端 ====================

class APIClient:
    """API 客户端基类。"""

    def __init__(self, config: Dict):
        self.config = config
        self.max_retries = int(config.get("max_retries", 3))
        self.timeout = int(config.get("timeout", 60))

    def call(self, prompt: str) -> str:
        raise NotImplementedError

    def test_connection(self) -> Tuple[bool, str]:
        raise NotImplementedError

    # -- 共用的带重试 POST --
    def _post_with_retry(self, url: str, headers: Dict, payload: Dict) -> requests.Response:
        last_exc: Optional[Exception] = None
        for attempt in range(self.max_retries):
            try:
                response = requests.post(url, headers=headers, json=payload, timeout=self.timeout)
                if response.status_code == 200:
                    return response
                msg = f"API 返回错误: HTTP {response.status_code} - {response.text[:200]}"
                last_exc = RuntimeError(msg)
                if attempt < self.max_retries - 1:
                    logger.warning("%s，重试 (%d/%d)", msg, attempt + 1, self.max_retries)
                    time.sleep(2 ** attempt)
                else:
                    raise last_exc
            except requests.Timeout as exc:
                last_exc = exc
                if attempt < self.max_retries - 1:
                    logger.warning("请求超时，重试 (%d/%d)", attempt + 1, self.max_retries)
                    time.sleep(2 ** attempt)
                else:
                    raise RuntimeError("API 请求超时，请检查网络或增大超时时间") from exc
            except requests.RequestException as exc:
                last_exc = exc
                if attempt < self.max_retries - 1:
                    logger.warning("请求失败: %s，重试 (%d/%d)", exc, attempt + 1, self.max_retries)
                    time.sleep(2 ** attempt)
                else:
                    raise RuntimeError(f"API 请求失败: {exc}") from exc
        # 理论不可达
        raise RuntimeError(f"API 调用失败: {last_exc}")


class OpenAICompatibleClient(APIClient):
    """OpenAI Chat Completions 兼容客户端。

    适用于 OpenAI、DeepSeek、智谱、通义千问、Azure、以及任何兼容
    ``/chat/completions`` 协议、用 ``Authorization: Bearer`` 鉴权的服务。
    """

    def _headers(self) -> Dict[str, str]:
        return {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.config['api_key']}",
        }

    def _payload(self, prompt: str, max_tokens: Optional[int] = None) -> Dict:
        return {
            "model": self.config["model"],
            "messages": [{"role": "user", "content": prompt}],
            "temperature": self.config.get("temperature", 0.0),
            "max_tokens": max_tokens or self.config.get("max_tokens", 8000),
        }

    def call(self, prompt: str) -> str:
        response = self._post_with_retry(self.config["api_url"], self._headers(), self._payload(prompt))
        result = response.json()
        try:
            return result["choices"][0]["message"]["content"]
        except (KeyError, IndexError, TypeError) as exc:
            raise RuntimeError(f"无法解析 API 响应结构: {result}") from exc

    def test_connection(self) -> Tuple[bool, str]:
        try:
            response = requests.post(
                self.config["api_url"],
                headers=self._headers(),
                json=self._payload("你好", max_tokens=50),
                timeout=10,
            )
            if response.status_code == 200:
                return True, "连接成功"
            return False, f"API 返回错误: HTTP {response.status_code} - {response.text[:200]}"
        except Exception as exc:
            return False, f"连接失败: {exc}"


class ClaudeClient(APIClient):
    """Anthropic Claude Messages API 客户端。"""

    def _headers(self) -> Dict[str, str]:
        return {
            "Content-Type": "application/json",
            "x-api-key": self.config["api_key"],
            "anthropic-version": "2023-06-01",
        }

    def _payload(self, prompt: str, max_tokens: Optional[int] = None) -> Dict:
        return {
            "model": self.config["model"],
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": max_tokens or self.config.get("max_tokens", 8000),
        }

    def call(self, prompt: str) -> str:
        response = self._post_with_retry(self.config["api_url"], self._headers(), self._payload(prompt))
        result = response.json()
        try:
            return result["content"][0]["text"]
        except (KeyError, IndexError, TypeError) as exc:
            raise RuntimeError(f"无法解析 Claude 响应结构: {result}") from exc

    def test_connection(self) -> Tuple[bool, str]:
        try:
            response = requests.post(
                self.config["api_url"],
                headers=self._headers(),
                json=self._payload("Hello", max_tokens=50),
                timeout=10,
            )
            if response.status_code == 200:
                return True, "连接成功"
            return False, f"API 返回错误: HTTP {response.status_code} - {response.text[:200]}"
        except Exception as exc:
            return False, f"连接失败: {exc}"


# provider 名称 -> 默认 api_url（api_url 未显式给出时使用）
PROVIDER_DEFAULT_URLS = {
    "DeepSeek": "https://api.deepseek.com/v1/chat/completions",
    "OpenAI": "https://api.openai.com/v1/chat/completions",
    "智谱AI": "https://open.bigmodel.cn/api/paas/v4/chat/completions",
    "通义千问": "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions",
}


def create_api_client(config: Dict) -> APIClient:
    """按 provider 创建对应的 API 客户端。

    ``config`` 需含 ``provider`` / ``model`` / ``api_key``，可选 ``api_url``、
    ``temperature`` / ``max_tokens`` / ``timeout`` / ``max_retries``。
    """
    config = dict(config)
    provider = config.get("provider", "DeepSeek")
    # 补全默认 url
    if not config.get("api_url"):
        config["api_url"] = PROVIDER_DEFAULT_URLS.get(provider, PROVIDER_DEFAULT_URLS["DeepSeek"])

    if provider == "Claude":
        return ClaudeClient(config)
    # DeepSeek / OpenAI / 智谱 / 通义 / Azure / 自定义 均走 OpenAI 兼容分支
    return OpenAICompatibleClient(config)


# ==================== 抽取引擎 ====================

class MedicalExtractionEngine:
    """把一段病历文本抽取为一行结构化数据。"""

    def __init__(self, api_config: Dict, template_config: Dict):
        self.template_config = template_config
        self.fields: List[Dict] = template_config.get("fields", [])
        if not self.fields:
            raise ValueError("模板配置缺少 fields，无法抽取")
        self.client = create_api_client(api_config)

    def build_prompt(self, emr_content: str) -> str:
        fields_info = []
        for field in self.fields:
            desc = field.get("description") or field["column"]
            fields_info.append(f"- {field['column']}: {desc} (类型: {field.get('type', '文本')})")
        fields_str = "\n".join(fields_info)
        emr_format = self.template_config.get("emr_format", "")

        return f"""你是一个专业的医疗数据提取助手。请从以下电子病历中提取信息，并按照指定的Excel模板字段填充数据。

电子病历格式说明：
{emr_format if emr_format else '标准电子病历格式'}

需要提取的字段：
{fields_str}

电子病历内容：
{emr_content}

请以JSON格式返回提取的数据，格式如下：
{{
    "字段1": "值1",
    "字段2": "值2"
}}

注意事项：
1. 严格按照字段的数据类型返回数据
2. 日期格式统一为 YYYY-MM-DD
3. 数字类型不要包含单位，只返回数字
4. 未提及的字段填 -1
5. 只返回JSON，不要包含其他解释文字
"""

    @staticmethod
    def parse_response(response: str) -> Dict:
        response = (response or "").strip()
        # 去掉 markdown ```json 代码块包裹
        if response.startswith("```"):
            lines = response.split("\n")
            if len(lines) > 2:
                response = "\n".join(lines[1:-1])
            if response.startswith("json"):
                response = response[4:].strip()
        try:
            return json.loads(response)
        except json.JSONDecodeError as exc:
            raise RuntimeError(f"解析 AI 响应失败: {exc}\n原始响应: {response[:200]}...") from exc

    def validate_data(self, data: Dict) -> Dict:
        validated: Dict = {}
        for field in self.fields:
            name = field["column"]
            ftype = field.get("type", "文本")
            value = data.get(name, "-1")
            try:
                if ftype in ("数字", "整数"):
                    validated[name] = -1 if value in ("-1", -1) else int(float(str(value)))
                elif ftype in ("小数", "浮点数"):
                    validated[name] = -1 if value in ("-1", -1) else float(str(value))
                else:
                    validated[name] = "-1" if value is None or value == "" else str(value)
            except (ValueError, TypeError):
                logger.warning("字段 %s 类型转换失败: %r -> %s", name, value, ftype)
                validated[name] = "-1"
        return validated

    def extract(self, emr_content: str, source: str = "") -> Dict:
        """抽取一份病历，返回带 ``_source`` / ``_status`` 的行数据。

        失败时抛异常由调用方决定是否隔离（不在此吞掉）。
        """
        prompt = self.build_prompt(emr_content)
        response = self.client.call(prompt)
        row = self.validate_data(self.parse_response(response))
        row["_source"] = source
        row["_status"] = "success"
        return row


# ==================== 模板配置 ====================

def load_template_config(source: str) -> Dict:
    """从预设 JSON 或 Excel 模板构造 template_config。

    - ``.json``：直接读取（需含 ``fields``，可含 ``emr_format``）。
    - ``.xlsx/.xls``：读取首行表头，每列生成一个「文本」字段。
    """
    path = Path(source)
    if not path.exists():
        raise FileNotFoundError(f"模板文件不存在: {source}")

    suffix = path.suffix.lower()
    if suffix == ".json":
        config = json.loads(path.read_text(encoding="utf-8"))
        if "fields" not in config:
            raise ValueError(f"模板 JSON 缺少 fields: {source}")
        return config

    if suffix in (".xlsx", ".xls"):
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        fields = [
            {"column": str(col).strip(), "description": "", "type": "文本"}
            for col in header
            if col is not None and str(col).strip()
        ]
        wb.close()
        if not fields:
            raise ValueError(f"Excel 模板首行没有可用表头: {source}")
        return {"template_path": str(path), "fields": fields, "emr_format": ""}

    raise ValueError(f"不支持的模板类型: {suffix}（请用 .json 或 .xlsx）")


# ==================== Excel 导出 ====================

def export_rows_to_excel(
    rows: List[Dict],
    template_path: str,
    output_path: str,
    log_callback: Optional[Callable[[str], None]] = None,
) -> str:
    """把抽取结果写入 Excel。

    以 ``template_path`` 为模板（保留其表头/格式），按表头列名匹配写入每行，
    并追加 ``提取状态`` / ``数据来源`` / ``错误信息`` 列，成功绿色、失败红色。
    返回实际写出的文件路径。
    """
    import openpyxl
    from openpyxl.styles import PatternFill

    def _log(msg: str):
        if log_callback:
            log_callback(msg)

    if not rows:
        raise ValueError("没有可导出的数据")

    tpl = Path(template_path)
    if not tpl.exists():
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    wb = openpyxl.load_workbook(str(tpl))
    ws = wb.active

    header_row = 1
    col_mapping: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(header_row, col).value
        if cell_value is not None and str(cell_value).strip():
            col_mapping[str(cell_value).strip()] = col

    # 追加元信息列
    for meta_key, meta_title in (("_status", "提取状态"), ("_source", "数据来源"), ("_error_message", "错误信息")):
        if meta_key not in col_mapping:
            new_col = ws.max_column + 1
            ws.cell(header_row, new_col, meta_title)
            col_mapping[meta_key] = new_col

    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for i, result in enumerate(rows):
        row_num = i + 2
        for field_name, col_idx in col_mapping.items():
            value = result.get(field_name, "")
            cell = ws.cell(row_num, col_idx, value)
            if field_name == "_status":
                if value == "success":
                    cell.fill = success_fill
                elif value == "failed":
                    cell.fill = failed_fill

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    _log(f"已导出 {len(rows)} 行到 {out}")
    return str(out)
