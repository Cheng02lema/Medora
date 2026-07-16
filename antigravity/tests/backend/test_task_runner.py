"""TaskRunner 测试：mock engine 后端，验证四态流转、continue-on-error、导出。"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch

from antigravity.backend.patient import PatientStore
from antigravity.backend.tasks import TaskRunner


def _updates_collector():
    updates = []
    return updates, (lambda p, m: updates.append((p.name, p.status, m)))


def test_full_chain_success(sample_parent, workspace, excel_template, tmp_path):
    store = PatientStore(workspace)
    store.import_parent(str(sample_parent))
    patients = store.all()

    out = tmp_path / "结果.xlsx"
    options = {
        "enable_preprocess": False,
        "enable_slice": False,
        "ocr_token": "tok",
        "ocr_url": "http://x",
        "ocr_model": "m",
        "extraction_template": str(excel_template),
        "extract_llm": {"provider": "DeepSeek", "api_key": "k", "model": "m"},
        "output_excel": str(out),
    }

    fake_ocr = MagicMock()
    fake_ocr.process_file.return_value = [
        {"layoutParsingResults": [{"markdown": {"text": "合成文本", "images": {}}, "outputImages": {}}]}
    ]
    fake_engine = MagicMock()
    fake_engine.extract.side_effect = lambda content, source="": {
        "姓名": source, "住院号": f"ID-{source}", "_source": source, "_status": "success"}

    updates, on_update = _updates_collector()
    runner = TaskRunner(patients, options, on_update=on_update)
    with patch("antigravity.engine.ocr_client.AsyncOCRClient", return_value=fake_ocr), \
         patch("antigravity.engine.medical_extractor.engine.MedicalExtractionEngine", return_value=fake_engine):
        summary = runner.run()

    assert summary["done"] == 2
    assert summary["error"] == 0
    assert summary["output_excel"]
    for p in patients:
        assert p.status == "done"
        assert p.row["姓名"] == p.name
    assert any(status == "running" for _, status, _ in updates)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    ncol = header.index("姓名") + 1
    names = {ws.cell(r, ncol).value for r in (2, 3)}
    assert names == {"张三", "李四"}


def test_continue_on_error_per_patient(sample_parent, workspace, excel_template):
    store = PatientStore(workspace)
    store.import_parent(str(sample_parent))
    patients = store.all()

    options = {
        "enable_preprocess": False,
        "enable_slice": False,
        "ocr_token": "tok",
        "extraction_template": str(excel_template),
        "extract_llm": {"provider": "DeepSeek", "api_key": "k", "model": "m"},
    }

    def _fake_client(*a, **k):
        m = MagicMock()

        def process_file(path):
            if "李四" in str(path):
                raise RuntimeError("OCR 崩了")
            return [{"layoutParsingResults": [{"markdown": {"text": "x", "images": {}}, "outputImages": {}}]}]

        m.process_file.side_effect = process_file
        return m

    fake_engine = MagicMock()
    fake_engine.extract.side_effect = lambda content, source="": {
        "姓名": source, "住院号": f"ID-{source}", "_source": source, "_status": "success"}

    runner = TaskRunner(patients, options)
    with patch("antigravity.engine.ocr_client.AsyncOCRClient", side_effect=_fake_client), \
         patch("antigravity.engine.medical_extractor.engine.MedicalExtractionEngine", return_value=fake_engine):
        summary = runner.run()

    states = {p.name: p.status for p in patients}
    assert states["张三"] == "done"
    assert states["李四"] == "error"
    assert "崩了" in patients[[p.name for p in patients].index("李四")].error_message


def test_missing_ocr_token_marks_error(sample_parent, workspace, excel_template):
    store = PatientStore(workspace)
    store.import_parent(str(sample_parent))
    patients = store.all()
    options = {"enable_preprocess": False, "enable_slice": False, "extraction_template": str(excel_template),
               "extract_llm": {"api_key": "k"}}
    runner = TaskRunner(patients, options)
    summary = runner.run()
    assert summary["error"] == 2
    for p in patients:
        assert p.status == "error"
        assert "OCR Token" in p.error_message


def test_stop_prevents_remaining_patients(sample_parent, workspace, excel_template):
    store = PatientStore(workspace)
    store.import_parent(str(sample_parent))
    patients = store.all()
    options = {"enable_preprocess": False, "enable_slice": False}
    runner = TaskRunner(patients, options)
    runner.stop()
    summary = runner.run()
    assert summary["done"] == 0 and summary["error"] == 0
    assert all(p.status == "pending" for p in patients)
